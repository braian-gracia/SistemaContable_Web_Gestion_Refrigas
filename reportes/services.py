from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.db.models import Sum, Count, Q
from django.utils import timezone
from cartera.models import Cliente, Deuda, Abono
from io import BytesIO
from decimal import Decimal

class ReporteService:
    
    @staticmethod
    def aplicar_estilos_encabezado(worksheet, ultima_columna):
        """Aplica estilos al encabezado del reporte"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for col in range(1, ultima_columna + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
    
    @staticmethod
    def ajustar_anchos_columna(worksheet):
        """Ajusta automáticamente el ancho de las columnas"""
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width
    
    @staticmethod
    def generar_reporte_general(fecha_inicio=None, fecha_fin=None, cliente_id=None, incluir_pagadas=False):
        """Genera un reporte general de deudas con filtros"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte General"
        
        # Encabezados
        headers = [
            'ID Deuda', 'Cliente', 'Correo Cliente', 'Teléfono', 'Descripción',
            'Monto Total', 'Total Abonado', 'Saldo Restante', 'Estado',
            'Fecha Creación', 'Fecha Vencimiento', 'Días Vencidos'
        ]
        ws.append(headers)
        
        # Filtrar deudas
        deudas = Deuda.objects.select_related('cliente').all()
        
        if fecha_inicio:
            deudas = deudas.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            deudas = deudas.filter(fecha__lte=fecha_fin)
        if cliente_id:
            deudas = deudas.filter(cliente_id=cliente_id)
        if not incluir_pagadas:
            deudas = deudas.filter(pagada=False)
        
        # Datos
        for deuda in deudas:
            saldo_restante = deuda.calcular_saldo_restante()
            total_abonado = deuda.monto - saldo_restante
            
            dias_vencidos = ''
            if deuda.fecha_vencimiento and deuda.esta_vencida():
                dias_vencidos = (timezone.now().date() - deuda.fecha_vencimiento).days
            
            estado = 'Pagada' if deuda.pagada else ('Vencida' if deuda.esta_vencida() else 'Pendiente')
            
            ws.append([
                deuda.id,
                deuda.cliente.nombre,
                deuda.cliente.correo,
                deuda.cliente.telefono or 'N/A',
                deuda.descripcion,
                float(deuda.monto),
                float(total_abonado),
                float(saldo_restante),
                estado,
                deuda.fecha.strftime('%Y-%m-%d'),
                deuda.fecha_vencimiento.strftime('%Y-%m-%d') if deuda.fecha_vencimiento else 'N/A',
                dias_vencidos
            ])
        
        # Totales
        ws.append([])
        totales_row = ws.max_row + 1
        ws.cell(row=totales_row, column=1, value='TOTALES')
        ws.cell(row=totales_row, column=1).font = Font(bold=True)
        
        total_monto = sum(float(d.monto) for d in deudas)
        total_saldo = sum(float(d.calcular_saldo_restante()) for d in deudas)
        total_abonado_general = total_monto - total_saldo
        
        ws.cell(row=totales_row, column=6, value=total_monto)
        ws.cell(row=totales_row, column=7, value=total_abonado_general)
        ws.cell(row=totales_row, column=8, value=total_saldo)
        
        # Estilos
        ReporteService.aplicar_estilos_encabezado(ws, len(headers))
        ReporteService.ajustar_anchos_columna(ws)
        
        # Guardar en memoria
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generar_reporte_clientes():
        """Genera un reporte detallado de todos los clientes"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Clientes"
        
        headers = [
            'ID Cliente', 'Nombre', 'Correo', 'Teléfono',
            'Total Deudas', 'Deudas Activas', 'Monto Total Adeudado',
            'Monto Total Pagado', 'Saldo Pendiente'
        ]
        ws.append(headers)
        
        clientes = Cliente.objects.all()
        
        for cliente in clientes:
            deudas = cliente.deuda_set.all()
            deudas_activas = deudas.filter(pagada=False)
            
            monto_total = sum(d.monto for d in deudas)
            saldo_pendiente = sum(d.calcular_saldo_restante() for d in deudas_activas)
            monto_pagado = monto_total - saldo_pendiente
            
            ws.append([
                cliente.id,
                cliente.nombre,
                cliente.correo,
                cliente.telefono or 'N/A',
                deudas.count(),
                deudas_activas.count(),
                float(monto_total),
                float(monto_pagado),
                float(saldo_pendiente)
            ])
        
        ReporteService.aplicar_estilos_encabezado(ws, len(headers))
        ReporteService.ajustar_anchos_columna(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer
    
    @staticmethod
    def generar_reporte_abonos(fecha_inicio=None, fecha_fin=None):
        """Genera un reporte de todos los abonos realizados"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte Abonos"
        
        headers = [
            'ID Abono', 'Cliente', 'Descripción Deuda', 'Monto Abono',
            'Fecha Abono', 'Saldo Restante Deuda', 'Estado Deuda'
        ]
        ws.append(headers)
        
        abonos = Abono.objects.select_related('deuda__cliente').all()
        
        if fecha_inicio:
            abonos = abonos.filter(fecha__gte=fecha_inicio)
        if fecha_fin:
            abonos = abonos.filter(fecha__lte=fecha_fin)
        
        for abono in abonos:
            ws.append([
                abono.id,
                abono.deuda.cliente.nombre,
                abono.deuda.descripcion,
                float(abono.monto),
                abono.fecha.strftime('%Y-%m-%d'),
                float(abono.deuda.calcular_saldo_restante()),
                'Pagada' if abono.deuda.pagada else 'Pendiente'
            ])
        
        # Total de abonos
        ws.append([])
        totales_row = ws.max_row + 1
        ws.cell(row=totales_row, column=1, value='TOTAL ABONOS')
        ws.cell(row=totales_row, column=1).font = Font(bold=True)
        ws.cell(row=totales_row, column=4, value=float(sum(a.monto for a in abonos)))
        
        ReporteService.aplicar_estilos_encabezado(ws, len(headers))
        ReporteService.ajustar_anchos_columna(ws)
        
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer